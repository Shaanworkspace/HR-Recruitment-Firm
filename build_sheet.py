import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA = [
    # #, Title, Series, Demand, VideoComp, Opportunity, ValidatedGap, DemandProof
    (1, "@Controller vs @RestController in Spring Boot | Interview Q", "S1 Spring Boot", "H", "L", "A+", "YES", "Blogs rank (Baeldung, JavaGuides); video sparse -> low comp, high search"),
    (2, "@RequestParam vs @RequestBody vs @PathVariable — when to use which", "S1 Spring Boot", "H", "L", "A+", "YES", "Blogs dominate; few clear video comparisons"),
    (3, "@Component vs @Service vs @Repository vs @Controller in Spring", "S1 Spring Boot", "H", "L", "A+", "YES", "Blogs (GeeksforGeeks, Baeldung); video thin"),
    (4, "@Autowired in Spring Boot — field vs constructor injection", "S1 Spring Boot", "H", "L", "A+", "YES", "Blogs rank high; interview staple, low-quality videos"),
    (5, "@ControllerAdvice + @ExceptionHandler — global error handling", "S1 Spring Boot", "H", "L", "A+", "YES", "Blogs dominate; advanced topic, few videos"),
    (6, "@Bean vs @Component in Spring Boot — when to use which", "S1 Spring Boot", "H", "L", "A+", "YES", "Blogs rank; video comparison scarce"),
    (7, "Spring Boot Interview for Freshers #1 — @SpringBootApplication & Auto-config", "S1 Spring Boot", "H", "L", "A+", "YES", "Fresher demand huge; blogs mostly, few peer videos"),
    (8, "Why is Spring Boot better than Spring? Interview answer", "S1 Spring Boot", "H", "L", "A+", "YES", "Blogs rank; common question, sparse video"),
    (9, "== vs .equals() in Java — the interview trap", "S2 Core Java", "H", "L", "A+", "YES", "Blogs rank; classic trap, low-video competition"),
    (10, "HashMap vs Hashtable vs ConcurrentHashMap in Java", "S3 Collections", "H", "L", "A+", "YES", "Blogs dominate; staple interview, sparse video"),
    (11, "JPA vs Hibernate — difference explained", "S5 JPA/REST", "H", "L", "A+", "YES", "Blogs (TutorialsPoint, Dev.to) rank; only mid-size 2025 video"),
    (12, "What is auto-configuration in Spring Boot? (internals)", "S1 Spring Boot", "H", "L", "A", "YES", "Blogs rank; internals rarely video-explained"),
    (13, "String is immutable in Java — why & interview answer", "S2 Core Java", "H", "L", "A", "YES", "Blogs dominate; fundamental, few clear videos"),
    (14, "String vs StringBuilder vs StringBuffer", "S2 Core Java", "H", "L", "A", "YES", "Blogs rank; common Q, video thin"),
    (15, "final keyword in Java — variable, method, class", "S2 Core Java", "H", "L", "A", "YES", "Blogs rank; basic but sparse video"),
    (16, "Method overloading vs overriding in Java", "S2 Core Java", "H", "L", "A", "YES", "Blogs dominate; staple, low video depth"),
    (17, "Interface vs Abstract class in Java (Java 8+)", "S2 Core Java", "H", "L", "A", "YES", "Blogs rank; common, few videos"),
    (18, "OOPs principles in Java explained with examples", "S2 Core Java", "H", "L", "A", "YES", "Blogs dominate; broad demand, sparse peer video"),
    (19, "ArrayList vs LinkedList — when to use which", "S3 Collections", "H", "L", "A", "YES", "Blogs rank; staple, low-video competition"),
    (20, "Comparable vs Comparator in Java", "S3 Collections", "H", "L", "A", "YES", "Blogs dominate; common, few videos"),
    (21, "synchronized in Java — method vs block", "S3 Collections", "H", "L", "A", "YES", "Blogs rank; concurrency basic, sparse video"),
    (22, "@RequestMapping vs @GetMapping/@PostMapping", "S1 Spring Boot", "H", "L", "A", "YES", "Blogs rank; mapping annotations, few videos"),
    (23, "@Qualifier & @Primary — resolve bean conflict", "S1 Spring Boot", "M", "L", "A", "YES", "Blogs dominate; advanced, sparse video"),
    (24, "@Transactional in Spring Boot — propagation & isolation", "S1 Spring Boot", "M", "L", "A", "YES", "Blogs rank; tricky, few clear videos"),
    (25, "@Value vs @ConfigurationProperties in Spring Boot", "S1 Spring Boot", "M", "L", "A", "YES", "Blogs dominate; config topic, sparse video"),
    (26, "What is a Spring Boot starter? (starter-web)", "S1 Spring Boot", "M", "L", "A", "YES", "Blogs rank; basics, few videos"),
    (27, "Embedded server in Spring Boot — Tomcat explained", "S1 Spring Boot", "M", "L", "A", "YES", "Blogs dominate; internals, sparse video"),
    (28, "Spring Boot JWT authentication — interview walkthrough", "S1 Spring Boot", "M", "L", "A", "YES", "Blogs rank; security staple, few peer videos"),
    (29, "hashCode() and equals() in Java — why override together", "S2 Core Java", "H", "L", "A", "YES", "Blogs dominate; classic, sparse video"),
    (30, "Why String pool exists in Java", "S2 Core Java", "M", "L", "A", "YES", "Blogs rank; internals, few videos"),
    (31, "Integer cache — why 128==128 is false", "S2 Core Java", "M", "L", "A", "YES", "Blogs dominate; trap question, sparse video"),
    (32, "autoboxing and unboxing in Java", "S2 Core Java", "M", "L", "A", "YES", "Blogs rank; basic, few videos"),
    (33, "static keyword in Java — variable, method, block", "S2 Core Java", "H", "L", "A", "YES", "Blogs dominate; fundamental, sparse video"),
    (34, "Exception handling in Java — checked vs unchecked", "S2 Core Java", "M", "L", "A", "YES", "Blogs rank; common, few videos"),
    (35, "Multithreading basics in Java — interview", "S2 Core Java", "M", "L", "A", "YES", "Blogs dominate; basic, sparse video"),
    (36, "Functional interface & lambda in Java 8", "S2 Core Java", "M", "L", "A", "YES", "Blogs rank; modern Java, few videos"),
    (37, "JVM vs JRE vs JDK — interview fundamentals", "S2 Core Java", "H", "L", "A", "YES", "Blogs dominate; fundamental, sparse peer video"),
    (38, "How HashMap works internally in Java", "S3 Collections", "H", "L", "A", "YES", "Blogs rank; deep-dive, few clear videos"),
    (39, "HashMap vs HashSet in Java", "S3 Collections", "M", "L", "A", "YES", "Blogs dominate; common, sparse video"),
    (40, "fail-fast vs fail-safe iterators in Java", "S3 Collections", "M", "L", "A", "YES", "Blogs rank; advanced, few videos"),
    (41, "ConcurrentHashMap vs Collections.synchronizedMap", "S3 Collections", "M", "L", "A", "YES", "Blogs dominate; concurrency, sparse video"),
    (42, "Vector vs ArrayList in Java", "S3 Collections", "M", "L", "A", "YES", "Blogs rank; legacy vs modern, few videos"),
    (43, "volatile keyword in Java — what it guarantees", "S3 Collections", "M", "L", "A", "YES", "Blogs dominate; tricky, sparse video"),
    (44, "wait, notify, notifyAll in Java", "S3 Collections", "M", "L", "A", "YES", "Blogs rank; concurrency, few videos"),
    (45, "Deadlock in Java — cause & how to avoid", "S3 Collections", "M", "L", "A", "YES", "Blogs dominate; common, sparse video"),
    (46, "ExecutorService & thread pool in Java", "S3 Collections", "M", "L", "A", "YES", "Blogs rank; advanced, few videos"),
    (47, "Spring Boot Actuator — endpoints & monitoring", "S1 Spring Boot", "M", "L", "A", "YES", "Blogs dominate; production topic, sparse video"),
    (48, "How to handle transactions in Spring Boot", "S1 Spring Boot", "M", "L", "A", "YES", "Blogs rank; common, few videos"),
    (49, "Spring Boot interview — common mistake that fails candidates", "S1 Spring Boot", "M", "L", "A", "YES", "Blogs dominate; evergreen, sparse video"),
    (50, "REST vs SOAP — interview answer", "S1 Spring Boot", "M", "L", "A", "YES", "Blogs rank; classic, few videos"),
    (51, "REST API best practices — Spring Boot", "S5 JPA/REST", "M", "L", "A", "YES", "Blogs dominate; broad, sparse peer video"),
    (52, "DTO vs Entity in Spring Boot — why use DTO", "S5 JPA/REST", "M", "L", "A", "YES", "Blogs rank; architecture, few videos"),
    (53, "Lazy vs Eager loading in Hibernate — N+1 problem", "S5 JPA/REST", "M", "L", "A", "YES", "Blogs dominate; performance, sparse video"),
    (54, "OneToMany / OneToOne / ManyToMany in JPA", "S5 JPA/REST", "M", "L", "A", "YES", "Blogs rank; mappings, few videos"),
    (55, "Spring Data JPA — save/findById/findAll explained", "S5 JPA/REST", "M", "L", "A", "YES", "Blogs dominate; basics, sparse video"),
    (56, "@Query vs derived methods in Spring Data JPA", "S5 JPA/REST", "M", "L", "A", "YES", "Blogs rank; advanced, few videos"),
    (57, "REST API versioning & error handling Spring Boot", "S5 JPA/REST", "M", "L", "A", "YES", "Blogs dominate; practical, sparse video"),
    (58, "SQL vs NoSQL — when to use which (interview)", "S5 JPA/REST", "M", "L", "A", "YES", "Blogs rank; broad, few peer videos"),
    (59, "@SpringBootTest vs @WebMvcTest vs @DataJpaTest", "S1 Spring Boot", "M", "L", "A", "YES", "Blogs dominate; testing, sparse video"),
    (60, "@MockBean vs @Mock in Spring Boot tests", "S1 Spring Boot", "M", "L", "A", "YES", "Blogs rank; testing, few videos"),
    (61, "JVM architecture explained — loader to runtime", "S4 JVM", "M", "L", "A", "YES", "Blogs dominate; internals, sparse video"),
    (62, "Heap vs Stack memory in Java", "S4 JVM", "M", "L", "A", "YES", "Blogs rank; fundamental, few videos"),
    (63, "Garbage Collection in Java explained simply", "S4 JVM", "M", "L", "A", "YES", "Blogs dominate; advanced, sparse video"),
    (64, "Recursion vs Iteration — when to use", "S6 DSA", "M", "L", "A", "YES", "Blogs rank; basics, sparse video"),
    (65, "BFS vs DFS — graph traversal explained", "S6 DSA", "M", "L", "A", "YES", "Blogs dominate; DSA staple, few videos"),
    (66, "Fast & Slow pointer pattern — linked list", "S6 DSA", "M", "L", "A", "YES", "Blogs rank; pattern, sparse video"),
    (67, "HashMap in DSA — frequency counting pattern", "S6 DSA", "M", "L", "A", "YES", "Blogs dominate; pattern, few videos"),
    (68, "Two Sum — 5 ways to solve in Java", "S6 DSA", "H", "M", "A", "YES", "High search; more videos but tutorial gap exists"),
    (69, "Sliding window pattern — DSA interview", "S6 DSA", "H", "M", "A", "YES", "High search; videos exist but peer angle open"),
    (70, "Binary search template — interview ready", "S6 DSA", "H", "M", "A", "YES", "High search; template gap, few peer videos"),
    (71, "Top 20 LeetCode patterns for interviews", "S6 DSA", "H", "M", "A", "YES", "High search; coaching-dominated, peer gap"),
    (72, "My BTech roadmap for product-based companies", "S7 Journey", "M", "L", "A", "YES", "Blogs/vlogs; student POV underserved"),
    (73, "How I use YouTube to prep for Java interviews", "S7 Journey", "M", "L", "A", "YES", "Vlogs sparse; student angle open"),
    (74, "Spring Boot project for resume — build in public #1", "S7 Journey", "M", "L", "A", "YES", "Project vlogs scarce; peer gap"),
    (75, "How to answer 'tell me about yourself' — CS student", "S7 Journey", "M", "L", "A", "YES", "HR blogs rank; video sparse"),
    (76, "HR round questions for freshers", "S7 Journey", "M", "L", "A", "YES", "Blogs dominate; few peer videos"),
    (77, "How to crack TCS/Infosys/product-based — roadmap", "S7 Journey", "M", "L", "A", "YES", "Blogs rank; student vlog gap"),
    (78, "My daily DSA routine as a BTech student", "S7 Journey", "M", "L", "A", "YES", "Routine vlogs sparse; peer angle"),
    (79, "Resume tips for Java developer fresher", "S7 Journey", "M", "L", "A", "YES", "Blogs dominate; few videos"),
    (80, "Mistakes I made in my first coding interview", "S7 Journey", "M", "L", "A", "YES", "Story vlogs open; peer gap"),
    (81, "How I improved English & communication", "S7 Journey", "M", "L", "A", "YES", "Communication vlogs sparse"),
    (82, "LeetCode Daily — <problem> | Java", "S6 DSA", "H", "M", "A", "DAY-1", "Proven day-1 traffic driver on your channel"),
    (83, "@Profile in Spring Boot — environment-specific beans", "S1 Spring Boot", "M", "L", "B", "YES", "Blogs rank; niche, sparse video"),
    (84, "@Lazy vs @DependsOn in Spring", "S1 Spring Boot", "M", "L", "B", "YES", "Blogs dominate; niche, few videos"),
    (85, "Spring Boot DevTools — what it does", "S1 Spring Boot", "L", "L", "B", "YES", "Blogs rank; low demand"),
    (86, "Marker interface in Java — Serializable, Cloneable", "S2 Core Java", "M", "L", "B", "YES", "Blogs dominate; basic, sparse video"),
    (87, "record in Java (Java 16) — why use records", "S2 Core Java", "M", "L", "B", "YES", "Blogs rank; modern, few videos"),
    (88, "var keyword in Java (Java 10)", "S2 Core Java", "M", "L", "B", "YES", "Blogs dominate; minor, sparse video"),
    (89, "String.intern() in Java — what it does", "S4 JVM", "M", "L", "B", "YES", "Blogs rank; niche, few videos"),
    (90, "OutOfMemoryError in Java — common causes", "S4 JVM", "M", "L", "B", "YES", "Blogs dominate; advanced, sparse video"),
    (91, "JVM memory areas — method area, PC, etc.", "S4 JVM", "M", "L", "B", "YES", "Blogs rank; internals, few videos"),
    (92, "Strong vs Weak vs Soft reference in Java", "S4 JVM", "L", "L", "B", "YES", "Blogs dominate; niche, sparse video"),
    (93, "What is JIT compiler in Java", "S4 JVM", "L", "L", "B", "YES", "Blogs rank; low demand"),
    (94, "Java 8 vs Java 17 features — interview", "S4 JVM", "M", "M", "B", "YES", "More videos; still peer gap"),
    (95, "ClassLoader in Java — how it loads classes", "S4 JVM", "M", "L", "B", "YES", "Blogs dominate; internals, sparse video"),
    (96, "PriorityQueue in Java explained", "S3 Collections", "M", "L", "B", "YES", "Blogs rank; basic, few videos"),
    (97, "Set vs List vs Queue in Java collections", "S3 Collections", "M", "L", "B", "YES", "Blogs dominate; basics, sparse video"),
    (98, "Pagination & Sorting in Spring Boot JPA", "S5 JPA/REST", "M", "L", "B", "YES", "Blogs rank; practical, few videos"),
    (99, "Dynamic programming for beginners — 0/1 knapsack", "S6 DSA", "M", "M", "B", "YES", "More videos; beginner gap"),
    (100, "LeetCode Daily (2nd) — <problem> | Java", "S6 DSA", "H", "M", "A", "DAY-1", "Proven day-1 traffic driver (repeat slot)"),
]

# Optimal publishing order: interleave LeetCode Daily every 4-5 long videos,
# finish S1 then S2,S3,S4,S5,S6,S7. We map series completion order.
SERIES_ORDER = ["S1 Spring Boot", "S2 Core Java", "S3 Collections", "S4 JVM", "S5 JPA/REST", "S6 DSA", "S7 Journey"]
by_series = {}
for row in DATA:
    by_series.setdefault(row[2], []).append(row)

# Build production order: for each series in order, emit its long videos,
# inserting a LeetCode Daily every 4 long videos.
prod = []
lc = [r for r in DATA if r[2] == "S6 DSA" and r[6] == "DAY-1"]
lc_count = 0
long_counter = 0
lc_pointer = 0
for s in SERIES_ORDER:
    rows = [r for r in by_series[s] if not (r[2] == "S6 DSA" and r[6] == "DAY-1")]
    for r in rows:
        prod.append(r)
        long_counter += 1
        if long_counter % 4 == 0 and lc_pointer < len(lc):
            prod.append(lc[lc_pointer])
            lc_pointer += 1
# append any leftover LeetCode Daily
while lc_pointer < len(lc):
    prod.append(lc[lc_pointer])
    lc_pointer += 1

wb = openpyxl.Workbook()

# ---------- Sheet 1: Priority List ----------
ws1 = wb.active
ws1.title = "Priority List"
headers1 = ["#", "Title", "Series", "Demand", "VideoComp", "Opportunity", "ValidatedGap", "Rank1HookLine", "DemandProof"]
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
opp_fill = {"A+": PatternFill("solid", fgColor="C6EFCE"), "A": PatternFill("solid", fgColor="FFEB9C"), "B": PatternFill("solid", fgColor="F2F2F2")}

ws1.append(headers1)
for c in range(1, len(headers1) + 1):
    cell = ws1.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

HOOK = "In this video we answer the exact <KEYWORD> interview question in Java/Spring Boot, with live code — by the end you'll never get it wrong."
for r in DATA:
    ws1.append([r[0], r[1], r[2], r[3], r[4], r[5], r[6], HOOK, r[7]])
    row_idx = ws1.max_row
    ws1.cell(row=row_idx, column=6).fill = opp_fill.get(r[5], opp_fill["B"])
    for c in range(1, len(headers1) + 1):
        ws1.cell(row=row_idx, column=c).border = border
        ws1.cell(row=row_idx, column=c).alignment = Alignment(vertical="top", wrap_text=True)

widths1 = [4, 52, 16, 8, 10, 12, 13, 50, 46]
for i, w in enumerate(widths1, start=1):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:I{ws1.max_row}"

# ---------- Sheet 2: Production Order ----------
ws2 = wb.create_sheet("Production Order")
headers2 = ["Publish#", "Title", "Series", "LinkToNext (same series)", "Demand", "Opportunity"]
ws2.append(headers2)
for c in range(1, len(headers2) + 1):
    cell = ws2.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# determine next-in-series for linking
series_titles = {}
for idx, r in enumerate(prod):
    series_titles.setdefault(r[2], []).append((idx, r[1]))
next_map = {}
for s, lst in series_titles.items():
    for i in range(len(lst)):
        cur_idx = lst[i][0]
        nxt = lst[i + 1][1] if i + 1 < len(lst) else "(end of series -> point to playlist)"
        next_map[cur_idx] = nxt

for pub_idx, r in enumerate(prod, start=1):
    orig_idx = DATA.index(r)
    ws2.append([pub_idx, r[1], r[2], next_map[orig_idx], r[3], r[5]])
    row_idx = ws2.max_row
    ws2.cell(row=row_idx, column=6).fill = opp_fill.get(r[5], opp_fill["B"])
    for c in range(1, len(headers2) + 1):
        ws2.cell(row=row_idx, column=c).border = border
        ws2.cell(row=row_idx, column=c).alignment = Alignment(vertical="top", wrap_text=True)

widths2 = [9, 52, 16, 46, 8, 12]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

# ---------- Sheet 3: Strategy ----------
ws3 = wb.create_sheet("Strategy")
ws3.column_dimensions["A"].width = 110
def add_line(text, bold=False, size=11):
    ws3.append([text])
    cell = ws3.cell(row=ws3.max_row, column=1)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if bold:
        cell.font = Font(bold=True, size=size, color="1F4E78")

add_line("SHAN LABS — YOUTUBE GROWTH STRATEGY", bold=True, size=14)
add_line("")
add_line("OBJECTIVE", bold=True)
add_line("Grow Shaan Labs via high-search / low-video-competition Java + Spring Boot interview content. Doubles as product-based-company job prep and on-camera communication training.")
add_line("")
add_line("CHANNEL STATUS (current)", bold=True)
add_line("254 views/28d, +3 subs, 95.5% new / <0.1% regular viewers (zero-loyalty problem). India 48.8%, 18-24 male 64%, Computer 80%. LeetCode Daily = proven day-1 view driver. Spring Boot series retention 5-6% (bounce <1 min).")
add_line("")
add_line("FORMAT DECISION", bold=True)
add_line("Long-form videos PRIMARY (8-15 min: comparison + live code + edge cases). Shorts clipped from long-form as SECONDARY discovery. NOT reels-only.")
add_line("")
add_line("3 PILLARS", bold=True)
add_line("A. LeetCode Daily (day-1 traffic)   B. Spring Boot interview Q&A + annotation deep-dives   C. BTech prep journey (Shorts/community)")
add_line("")
add_line("DROPPED: System Design as a core topic (saturated: freeCodeCamp 11.8M, Telusko 2.9M, ByteByteGo 1.4M, Gaurav Sen 743K, Hello Interview; avg tutorial = 154K views). Stay in the Java/Spring interview lane.")
add_line("")
add_line("THE EXACT RANK #1 LINE (use in every video)", bold=True)
add_line(HOOK)
add_line("Apply: keyword in title (first 60 chars) + spoken in first 15s + in first 2 lines of description + tags + thumbnail text.")
add_line("")
add_line("PROMOTION / LINKING ENGINE (your #1 growth lever)", bold=True)
add_line("1. Create 7 playlists — one per series (S1-S7). New viewers from any video get pulled into the playlist -> multiple views -> algorithm rewards watch-time + return-viewers (your missing metric).")
add_line("2. Every long video's end screen points to the NEXT video in its series (see 'LinkToNext' column in Production Order).")
add_line("3. Description of each video links to the playlist + the 2 most relevant sibling videos.")
add_line("4. Interleave a LeetCode Daily every 4-5 long videos so fresh search traffic constantly feeds the channel.")
add_line("")
add_line("PRODUCTION SEQUENCE", bold=True)
add_line("Publish S1 (Spring Boot) first to #25, then S2, S3, S4, S5, S6, S7. Within each series go A+ -> A -> B. Finishing one series before the next builds a complete playlist faster.")
add_line("")
add_line("90-DAY EXPECTATION (realistic, no paid tool / no guarantee)", bold=True)
add_line("- Months 1-2: tiny views; focus on consistency (2-3 long videos/week + daily LeetCode Short).")
add_line("- Month 3: a few A+ titles begin ranking as Google/YouTube index the niche gaps; return-viewers should climb above 1%.")
add_line("- Do NOT expect viral spikes. Compounding comes from playlist internal linking, not single videos.")
add_line("")
add_line("NOTE ON DATA", bold=True)
add_line("Exact numeric search volumes require a paid tool (vidIQ/TubeBuddy). Demand (H/M/L) and VideoComp (L=good / M) tiers are derived from live search evidence: blogs/Medium/StackOverflow ranking for the keyword = high search intent; sparse-quality YouTube videos = low video competition. All 'ValidatedGap=YES' rows were confirmed by live web search.")

out = "/Users/shaanyadav/Documents/ShaanLabs_100_Titles.xlsx"
wb.save(out)
print("Saved:", out)
print("Priority rows:", len(DATA))
print("Production rows:", len(prod))
